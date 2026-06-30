"""
scripts/curva_abc.py
====================
Extração de relatórios "Curva ABC de Vendas de Produtos" (sistema RADInfo)
e geração do Excel mensal.

Esta é a LÓGICA TESTADA original (curva_abc_para_excel.py), convertida em
módulo importável. As funções de extração por coordenadas (parse_page,
extrair_dados) e a geração do Excel (gerar_excel) foram mantidas intactas.
A interface de linha de comando foi removida — quem orquestra agora é
scripts/relatorios_vendas.py.

Dependências:
    pip install pdfplumber openpyxl
"""

import re

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Limites das colunas (coordenada X em pontos)
# Baseado no layout padrão do relatório RADInfo
# ---------------------------------------------------------------------------
COL_BOUNDS = {
    'interv': (0,   47),
    'codigo': (47,  91),
    'barras': (91,  181),
    'desc':   (181, 411),
    'qtd':    (411, 441),
    'valor':  (441, 513),
    'perc':   (513, 552),
    'classe': (552, 600),
}

VALOR_PAT = re.compile(r'\d{1,3}(?:\.\d{3})*,\d{2}')
PERC_PAT  = re.compile(r'\d+,\d+')


def col_name(x0: float) -> str:
    for name, (lo, hi) in COL_BOUNDS.items():
        if lo <= x0 < hi:
            return name
    return 'outro'


def br_float(s: str) -> float:
    return float(s.replace('.', '').replace(',', '.'))


# ---------------------------------------------------------------------------
# Extração por coordenadas
# ---------------------------------------------------------------------------

def parse_page(page) -> list[dict]:
    words = page.extract_words(x_tolerance=3, y_tolerance=3)

    # Agrupa palavras por linha (tolerância de 2pt no eixo Y)
    rows: dict[int, list] = {}
    for w in words:
        key = round(w['top'])
        rows.setdefault(key, []).append(w)

    produtos = []
    for top in sorted(rows):
        ws = sorted(rows[top], key=lambda w: w['x0'])
        if not ws:
            continue

        first = ws[0]
        # Linha de produto começa com inteiro na coluna Interv
        if col_name(first['x0']) != 'interv':
            continue
        if not re.match(r'^\d+$', first['text']):
            continue

        # Distribui tokens nas colunas
        buckets: dict[str, list[str]] = {c: [] for c in COL_BOUNDS}
        for w in ws:
            c = col_name(w['x0'])
            if c in buckets:
                buckets[c].append(w['text'])

        interv = buckets['interv'][0] if buckets['interv'] else ''
        codigo = buckets['codigo'][0] if buckets['codigo'] else ''
        barras = buckets['barras'][0] if buckets['barras'] else ''
        desc   = ' '.join(buckets['desc'])

        # Qtd: apenas tokens puramente numéricos na coluna qtd
        qtd_tokens = [t for t in buckets['qtd'] if re.match(r'^\d+$', t)]
        qtd = int(qtd_tokens[0]) if qtd_tokens else 0

        # Valor: extrai padrão monetário dos tokens da coluna valor
        valor_raw = ' '.join(buckets['valor'])
        mv = VALOR_PAT.findall(valor_raw)
        valor = br_float(mv[-1]) if mv else 0.0

        # Perc: primeiro nn,nn na coluna perc
        perc_raw = ' '.join(buckets['perc'])
        mp = PERC_PAT.search(perc_raw)
        perc = float(mp.group().replace(',', '.')) if mp else 0.0

        # Classe
        classe = buckets['classe'][0] if buckets['classe'] else '?'
        # Garante que seja apenas A, B ou C
        if classe not in ('A', 'B', 'C'):
            m = re.search(r'[ABC]', classe)
            classe = m.group() if m else '?'

        if interv and codigo:
            produtos.append({
                'Intervalo':        int(interv),
                'Código':           codigo,
                'Código de Barras': barras,
                'Descrição':        desc,
                'Qtd':              qtd,
                'Valor Total':      valor,
                'Perc %':           perc,
                'Classe':           classe,
            })

    return produtos


def extrair_dados(pdf_path: str) -> dict:
    produtos = []
    totais_classe = {}
    total_geral = None
    periodo = ''
    loja = ''

    RE_PERIODO     = re.compile(r'(\d{2}/\d{2}/\d{4})\s+a\s+(\d{2}/\d{2}/\d{4})')
    RE_LOJA        = re.compile(r'SUPERMERCADO\s+\S+\s+\S+', re.IGNORECASE)
    RE_TOT_CLASSE  = re.compile(r'Total Valor de Produtos Classe ([ABC]):\s+([\d.,]+)\s+([\d,]+)\s*%')
    RE_TOT_GERAL   = re.compile(r'Total Geral:\s+([\d.,]+)')

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            # Metadados e totais via texto simples
            texto = page.extract_text() or ''
            if not periodo:
                m = RE_PERIODO.search(texto)
                if m:
                    periodo = f'{m.group(1)} a {m.group(2)}'
            if not loja:
                m = RE_LOJA.search(texto)
                if m:
                    loja = m.group()
            for m in RE_TOT_CLASSE.finditer(texto):
                totais_classe[m.group(1)] = {
                    'valor': br_float(m.group(2)),
                    'perc':  br_float(m.group(3)),
                }
            m = RE_TOT_GERAL.search(texto)
            if m:
                total_geral = br_float(m.group(1))

            # Produtos via coordenadas
            produtos.extend(parse_page(page))

    # Ordena por intervalo para garantir sequência
    produtos.sort(key=lambda p: p['Intervalo'])

    return {
        'produtos':      produtos,
        'totais_classe': totais_classe,
        'total_geral':   total_geral,
        'periodo':       periodo,
        'loja':          loja,
    }


# ---------------------------------------------------------------------------
# Estilos Excel
# ---------------------------------------------------------------------------

COR_HEADER   = '1F4E79'
COR_CLASSE_A = 'D9EAD3'
COR_CLASSE_B = 'FFF2CC'
COR_CLASSE_C = 'FCE4D6'
COR_TOTAL    = 'D9D9D9'

BORDA = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin'),
)


def aplicar_header(ws, cols: list[str]):
    ws.append(cols)
    r = ws.max_row
    for c in range(1, len(cols) + 1):
        cell = ws.cell(r, c)
        cell.font      = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        cell.fill      = PatternFill('solid', start_color=COR_HEADER)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = BORDA
    ws.row_dimensions[r].height = 30


def cor_classe(classe: str) -> str:
    return {'A': COR_CLASSE_A, 'B': COR_CLASSE_B, 'C': COR_CLASSE_C}.get(classe, 'FFFFFF')


# ---------------------------------------------------------------------------
# Geração do Excel mensal
# ---------------------------------------------------------------------------

def gerar_excel(dados: dict, saida: str):
    wb = Workbook()

    # ── Aba 1: Curva ABC ─────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Curva ABC'

    cols = ['Intervalo', 'Código', 'Código de Barras', 'Descrição',
            'Qtd', 'Valor Total (R$)', 'Perc %', 'Classe']
    aplicar_header(ws, cols)

    for prod in dados['produtos']:
        ws.append([
            prod['Intervalo'],
            prod['Código'],
            prod['Código de Barras'],
            prod['Descrição'],
            prod['Qtd'],
            prod['Valor Total'],
            prod['Perc %'] / 100,
            prod['Classe'],
        ])
        r = ws.max_row
        fill = PatternFill('solid', start_color=cor_classe(prod['Classe']))
        for c in range(1, len(cols) + 1):
            cell = ws.cell(r, c)
            cell.fill      = fill
            cell.border    = BORDA
            cell.font      = Font(name='Arial', size=9)
            cell.alignment = Alignment(vertical='center')
        ws.cell(r, 1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(r, 5).alignment = Alignment(horizontal='right',  vertical='center')
        ws.cell(r, 6).alignment = Alignment(horizontal='right',  vertical='center')
        ws.cell(r, 7).alignment = Alignment(horizontal='right',  vertical='center')
        ws.cell(r, 8).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(r, 5).number_format = '#,##0'
        ws.cell(r, 6).number_format = '#,##0.00'
        ws.cell(r, 7).number_format = '0.00%'

    # Linha de total
    last = ws.max_row + 1
    ws.cell(last, 4).value = 'TOTAL GERAL'
    ws.cell(last, 5).value = f'=SUM(E2:E{last-1})'
    ws.cell(last, 5).number_format = '#,##0'
    ws.cell(last, 6).value = f'=SUM(F2:F{last-1})'
    ws.cell(last, 6).number_format = '#,##0.00'
    ws.cell(last, 7).value = f'=SUM(G2:G{last-1})'
    ws.cell(last, 7).number_format = '0.00%'
    for c in range(1, len(cols) + 1):
        cell = ws.cell(last, c)
        cell.fill   = PatternFill('solid', start_color=COR_TOTAL)
        cell.border = BORDA
        cell.font   = Font(bold=True, name='Arial', size=9)

    for i, w in enumerate([10, 10, 22, 62, 10, 18, 10, 8], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    # ── Aba 2: Resumo por Classe ─────────────────────────────────────────
    ws2 = wb.create_sheet('Resumo por Classe')
    ws2['A1'] = dados.get('loja', '')
    ws2['A1'].font = Font(bold=True, name='Arial', size=11)
    ws2['A2'] = f"Período: {dados.get('periodo', '')}"
    ws2['A2'].font = Font(name='Arial', size=10, italic=True)
    ws2.append([])
    aplicar_header(ws2, ['Classe', 'Qtd Produtos', 'Valor Total (R$)', 'Participação %'])

    for classe in ['A', 'B', 'C']:
        prods = [p for p in dados['produtos'] if p['Classe'] == classe]
        valor = dados['totais_classe'].get(classe, {}).get('valor',
                    sum(p['Valor Total'] for p in prods))
        perc  = dados['totais_classe'].get(classe, {}).get('perc', 0)
        ws2.append([classe, len(prods), valor, perc / 100])
        r = ws2.max_row
        fill = PatternFill('solid', start_color=cor_classe(classe))
        for c in range(1, 5):
            ws2.cell(r, c).fill   = fill
            ws2.cell(r, c).border = BORDA
            ws2.cell(r, c).font   = Font(name='Arial', size=10)
        ws2.cell(r, 3).number_format = '#,##0.00'
        ws2.cell(r, 4).number_format = '0.00%'

    t = ws2.max_row + 1
    ws2.cell(t, 1).value = 'TOTAL'
    ws2.cell(t, 2).value = len(dados['produtos'])
    ws2.cell(t, 3).value = dados['total_geral'] or sum(p['Valor Total'] for p in dados['produtos'])
    ws2.cell(t, 4).value = 1.0
    for c in range(1, 5):
        ws2.cell(t, c).fill   = PatternFill('solid', start_color=COR_TOTAL)
        ws2.cell(t, c).border = BORDA
        ws2.cell(t, c).font   = Font(bold=True, name='Arial', size=10)
    ws2.cell(t, 3).number_format = '#,##0.00'
    ws2.cell(t, 4).number_format = '0.00%'
    for i, w in zip(range(1, 5), [10, 15, 20, 16]):
        ws2.column_dimensions[get_column_letter(i)].width = w

    wb.save(saida)
