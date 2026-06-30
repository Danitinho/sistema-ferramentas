"""
scripts/relatorios_vendas.py
============================
Módulo de Relatórios de Venda (Curva ABC).

Responsabilidades:
  • Ler PDFs da pasta de ENTRADA, extrair os dados (lógica testada em
    scripts/curva_abc.py), gerar o Excel mensal na pasta de SAÍDA e gravar
    os dados consolidados no banco SQLite.
  • Mover o PDF processado para a pasta de PROCESSADOS.
  • Consultar uma lista de códigos de barras e retornar, para cada um,
    o código, a descrição (mais recente) e a quantidade vendida por mês.

Por que SQLite e não uma pasta de planilhas?
  A consulta "quanto cada produto vendeu mês a mês" precisa cruzar todos os
  meses. Com planilhas soltas seria preciso reabrir e reparsear cada arquivo
  a cada consulta. No SQLite isso vira uma única query indexada — instantânea
  e estável conforme os meses se acumulam. Os Excel mensais continuam sendo
  gerados normalmente (para leitura humana); o banco é só a fonte de consulta.

Estrutura de pastas (criadas automaticamente dentro de dados/relatorios/):
    entrada/      → solte aqui os PDFs novos
    processados/  → PDFs já lidos vão para cá
    saida/        → Excel mensal de cada relatório
    vendas.db     → banco consolidado
"""

import os
import re
import sqlite3
import shutil
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from scripts import curva_abc

# ── Caminhos ─────────────────────────────────────────────────────────────────
BASE_DIR      = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAIZ          = os.path.join(BASE_DIR, "dados", "relatorios")
DIR_ENTRADA   = os.path.join(RAIZ, "entrada")
DIR_PROCESSADO = os.path.join(RAIZ, "processados")
DIR_SAIDA     = os.path.join(RAIZ, "saida")
BANCO         = os.path.join(RAIZ, "vendas.db")

MESES_PT = ["", "jan", "fev", "mar", "abr", "mai", "jun",
            "jul", "ago", "set", "out", "nov", "dez"]


# ── Infraestrutura ───────────────────────────────────────────────────────────
def _garantir_pastas():
    for d in (DIR_ENTRADA, DIR_PROCESSADO, DIR_SAIDA):
        os.makedirs(d, exist_ok=True)


def _conn():
    _garantir_pastas()
    conn = sqlite3.connect(BANCO)
    conn.row_factory = sqlite3.Row
    _init_schema(conn)
    return conn


def _init_schema(conn):
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS relatorios (
            mes          TEXT PRIMARY KEY,   -- 'YYYY-MM'
            periodo      TEXT,               -- '01/05/2025 a 31/05/2025'
            loja         TEXT,
            total_geral  REAL,
            arquivo      TEXT,
            processado_em TEXT
        );
        CREATE TABLE IF NOT EXISTS vendas (
            mes           TEXT NOT NULL,
            intervalo     INTEGER,
            codigo        TEXT,
            codigo_barras TEXT,
            descricao     TEXT,
            qtd           REAL,
            valor_total   REAL,
            perc          REAL,
            classe        TEXT,
            FOREIGN KEY (mes) REFERENCES relatorios(mes) ON DELETE CASCADE
        );
        CREATE INDEX IF NOT EXISTS idx_vendas_barras ON vendas(codigo_barras);
        CREATE INDEX IF NOT EXISTS idx_vendas_codigo ON vendas(codigo);
        CREATE INDEX IF NOT EXISTS idx_vendas_mes    ON vendas(mes);
    """)
    conn.commit()
    _migrar_qtd_para_real(conn)


def _migrar_qtd_para_real(conn):
    """Migra qtd de INTEGER para REAL caso o banco seja de uma versão anterior."""
    info = conn.execute("PRAGMA table_info(vendas)").fetchall()
    col = next((r for r in info if r["name"] == "qtd"), None)
    if col and col["type"].upper() == "INTEGER":
        conn.executescript("""
            ALTER TABLE vendas RENAME TO _vendas_old;
            CREATE TABLE vendas (
                mes           TEXT NOT NULL,
                intervalo     INTEGER,
                codigo        TEXT,
                codigo_barras TEXT,
                descricao     TEXT,
                qtd           REAL,
                valor_total   REAL,
                perc          REAL,
                classe        TEXT,
                FOREIGN KEY (mes) REFERENCES relatorios(mes) ON DELETE CASCADE
            );
            INSERT INTO vendas SELECT * FROM _vendas_old;
            DROP TABLE _vendas_old;
            CREATE INDEX IF NOT EXISTS idx_vendas_barras ON vendas(codigo_barras);
            CREATE INDEX IF NOT EXISTS idx_vendas_codigo ON vendas(codigo);
            CREATE INDEX IF NOT EXISTS idx_vendas_mes    ON vendas(mes);
        """)
        conn.commit()


# ── Utilitários de mês ───────────────────────────────────────────────────────
def _mes_de_periodo(periodo: str) -> str | None:
    """'01/05/2025 a 31/05/2025' → '2025-05' (usa a data inicial)."""
    m = re.search(r'(\d{2})/(\d{2})/(\d{4})', periodo or "")
    if not m:
        return None
    _dia, mes, ano = m.groups()
    return f"{ano}-{mes}"


def rotulo_mes(mes: str) -> str:
    """'2025-05' → 'mai/2025' (rótulo amigável para a tela e o Excel)."""
    try:
        ano, m = mes.split("-")
        return f"{MESES_PT[int(m)]}/{ano}"
    except (ValueError, IndexError):
        return mes


# ── Gravação no banco ────────────────────────────────────────────────────────
def _gravar_no_banco(conn, mes: str, dados: dict, arquivo: str):
    """Substitui por completo os dados do mês (reprocessar é seguro)."""
    conn.execute("DELETE FROM vendas     WHERE mes = ?", (mes,))
    conn.execute("DELETE FROM relatorios WHERE mes = ?", (mes,))

    conn.execute(
        "INSERT INTO relatorios (mes, periodo, loja, total_geral, arquivo, processado_em) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        (mes, dados.get("periodo", ""), dados.get("loja", ""),
         dados.get("total_geral"), arquivo,
         datetime.now().strftime("%d/%m/%Y %H:%M")),
    )

    conn.executemany(
        "INSERT INTO vendas (mes, intervalo, codigo, codigo_barras, descricao, "
        "qtd, valor_total, perc, classe) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
        [
            (mes, p["Intervalo"], p["Código"], p["Código de Barras"],
             p["Descrição"], p["Qtd"], p["Valor Total"], p["Perc %"], p["Classe"])
            for p in dados["produtos"]
        ],
    )
    conn.commit()


# ── Processamento da pasta de entrada ────────────────────────────────────────
def processar_entrada() -> dict:
    """
    Lê todos os PDFs de DIR_ENTRADA. Para cada um: extrai, gera o Excel mensal,
    grava no banco e move o PDF para DIR_PROCESSADO. PDFs com erro NÃO são
    movidos (ficam na entrada para inspeção).

    Retorna um relatório do que aconteceu.
    """
    _garantir_pastas()
    pdfs = sorted(
        f for f in os.listdir(DIR_ENTRADA)
        if f.lower().endswith(".pdf")
    )

    resultados = []
    conn = _conn()
    try:
        for nome in pdfs:
            caminho = os.path.join(DIR_ENTRADA, nome)
            try:
                dados = curva_abc.extrair_dados(caminho)

                if not dados["produtos"]:
                    resultados.append({
                        "arquivo": nome, "ok": False,
                        "msg": "Nenhum produto extraído (PDF fora do layout esperado?)."
                    })
                    continue

                mes = _mes_de_periodo(dados.get("periodo", ""))
                if not mes:
                    resultados.append({
                        "arquivo": nome, "ok": False,
                        "msg": "Não foi possível identificar o período/mês no PDF."
                    })
                    continue

                # Excel mensal
                nome_excel = f"curva_abc_{mes}.xlsx"
                caminho_excel = os.path.join(DIR_SAIDA, nome_excel)
                curva_abc.gerar_excel(dados, caminho_excel)

                # Banco (substitui o mês)
                ja_existia = conn.execute(
                    "SELECT 1 FROM relatorios WHERE mes = ?", (mes,)
                ).fetchone() is not None
                _gravar_no_banco(conn, mes, dados, nome)

                # Move o PDF para processados (sobrescreve se já houver homônimo)
                destino = os.path.join(DIR_PROCESSADO, nome)
                if os.path.exists(destino):
                    os.remove(destino)
                shutil.move(caminho, destino)

                resultados.append({
                    "arquivo": nome, "ok": True,
                    "mes": mes, "rotulo": rotulo_mes(mes),
                    "produtos": len(dados["produtos"]),
                    "excel": nome_excel,
                    "substituido": ja_existia,
                    "msg": (f"{len(dados['produtos'])} produtos · "
                            f"{rotulo_mes(mes)}"
                            + (" (mês substituído)" if ja_existia else "")),
                })
            except Exception as e:  # noqa: BLE001 — queremos continuar nos demais
                resultados.append({
                    "arquivo": nome, "ok": False,
                    "msg": f"Erro ao processar: {e}"
                })
    finally:
        conn.close()

    ok = [r for r in resultados if r["ok"]]
    return {
        "total":      len(resultados),
        "processados": len(ok),
        "falhas":     len(resultados) - len(ok),
        "itens":      resultados,
    }


# ── Estado / metadados ───────────────────────────────────────────────────────
def status() -> dict:
    """Resumo para exibir na tela: meses no banco, total de produtos, fila."""
    _garantir_pastas()
    conn = _conn()
    try:
        meses = [
            {"mes": r["mes"], "rotulo": rotulo_mes(r["mes"]),
             "periodo": r["periodo"], "produtos": _qtd_produtos(conn, r["mes"])}
            for r in conn.execute(
                "SELECT mes, periodo FROM relatorios ORDER BY mes").fetchall()
        ]
        total_linhas = conn.execute("SELECT COUNT(*) FROM vendas").fetchone()[0]
    finally:
        conn.close()

    pendentes = [f for f in os.listdir(DIR_ENTRADA) if f.lower().endswith(".pdf")] \
        if os.path.isdir(DIR_ENTRADA) else []

    return {
        "meses":          meses,
        "total_meses":    len(meses),
        "total_registros": total_linhas,
        "pendentes":      sorted(pendentes),
        "pasta_entrada":  DIR_ENTRADA,
    }


def _qtd_produtos(conn, mes):
    return conn.execute("SELECT COUNT(*) FROM vendas WHERE mes = ?", (mes,)).fetchone()[0]


def _meses_ordenados(conn) -> list[str]:
    return [r["mes"] for r in conn.execute(
        "SELECT mes FROM relatorios ORDER BY mes").fetchall()]


# ── Consulta por código de barras ────────────────────────────────────────────
def _normaliza_codigos(texto_ou_lista) -> list[str]:
    """Aceita string (várias linhas / vírgulas) ou lista. Remove vazios/duplicados."""
    if isinstance(texto_ou_lista, str):
        brutos = re.split(r'[\s,;]+', texto_ou_lista)
    else:
        brutos = []
        for item in (texto_ou_lista or []):
            brutos.extend(re.split(r'[\s,;]+', str(item)))
    vistos, saida = set(), []
    for c in brutos:
        c = c.strip()
        if c and c not in vistos:
            vistos.add(c)
            saida.append(c)
    return saida


def consultar_codigos(codigos) -> dict:
    """
    Para cada código de barras informado, retorna:
      codigo (interno), descrição (a mais recente entre os meses),
      e a quantidade vendida em cada mês registrado.

    Retorno:
      {
        "meses":  ["2025-04", "2025-05", ...],
        "rotulos": ["abr/2025", "mai/2025", ...],
        "linhas": [
          {"codigo_barras": "...", "codigo": "...", "descricao": "...",
           "encontrado": True, "qtds": {"2025-04": 12, "2025-05": 30},
           "total": 42},
          ...
        ]
      }
    """
    codigos = _normaliza_codigos(codigos)
    conn = _conn()
    try:
        meses = _meses_ordenados(conn)
        linhas = []
        for cb in codigos:
            regs = conn.execute(
                "SELECT mes, codigo, descricao, qtd FROM vendas "
                "WHERE codigo_barras = ? ORDER BY mes",
                (cb,),
            ).fetchall()

            if not regs:
                linhas.append({
                    "codigo_barras": cb, "codigo": "", "descricao": "",
                    "encontrado": False, "qtds": {}, "total": 0,
                })
                continue

            qtds = {r["mes"]: (r["qtd"] or 0) for r in regs}
            # descrição/código mais recentes = último mês com registro
            ultimo = regs[-1]
            linhas.append({
                "codigo_barras": cb,
                "codigo":        ultimo["codigo"] or "",
                "descricao":     ultimo["descricao"] or "",
                "encontrado":    True,
                "qtds":          qtds,
                "total":         sum(qtds.values()),
            })
    finally:
        conn.close()

    return {
        "meses":   meses,
        "rotulos": [rotulo_mes(m) for m in meses],
        "linhas":  linhas,
    }


# ── Excel da consulta ────────────────────────────────────────────────────────
_COR_HEADER = "1F4E79"
_COR_ZEBRA  = "F1EFE8"
_COR_NAO    = "FCEBEB"
_BORDA = Border(left=Side("thin"), right=Side("thin"),
                top=Side("thin"), bottom=Side("thin"))


def gerar_excel_consulta(resultado: dict, caminho: str):
    """Gera o Excel pivotado da consulta (uma coluna por mês)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Consulta"

    rotulos = resultado["rotulos"]
    meses   = resultado["meses"]
    cols = ["Código de Barras", "Código", "Descrição", *rotulos, "Total"]

    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(1, c)
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill      = PatternFill("solid", start_color=_COR_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _BORDA
    ws.row_dimensions[1].height = 28

    for i, ln in enumerate(resultado["linhas"]):
        linha = [ln["codigo_barras"], ln["codigo"], ln["descricao"]]
        linha += [ln["qtds"].get(m, 0) for m in meses]
        linha.append(ln["total"])
        ws.append(linha)
        r = ws.max_row

        if not ln["encontrado"]:
            fill = PatternFill("solid", start_color=_COR_NAO)
            ws.cell(r, 3).value = "— não encontrado —"
        else:
            fill = PatternFill("solid", start_color=_COR_ZEBRA) if i % 2 else None

        for c in range(1, len(cols) + 1):
            cell = ws.cell(r, c)
            cell.font   = Font(name="Arial", size=9)
            cell.border = _BORDA
            if fill:
                cell.fill = fill
            if c >= 4:  # colunas numéricas
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0.###"

    larguras = [22, 12, 50] + [12] * len(meses) + [12]
    for i, w in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "D2"

    wb.save(caminho)
