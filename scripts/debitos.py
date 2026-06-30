"""
scripts/debitos.py
Módulo de débitos — vencimentos e rebaixas de preço.
Armazena os dados em dados/debitos.xlsx (criado automaticamente).
"""
import os
import uuid
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── Caminhos ─────────────────────────────────────────────────────────────────
BASE_DIR    = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DADOS_DIR   = os.path.join(BASE_DIR, "dados")
ARQUIVO     = os.path.join(DADOS_DIR, "debitos.xlsx")

# ── Cabeçalhos das abas ───────────────────────────────────────────────────────
H_EMPRESAS      = ["cnpj", "razao_social"]
H_DEBITOS       = ["id", "data", "cnpj", "razao_social", "tipo",
                   "nf_numero", "produto", "quantidade", "valor_unit", "valor_debito", "obs"]
H_BONIFICACOES  = ["id", "data", "cnpj", "razao_social", "nf_numero",
                   "produto", "quantidade", "valor_unit", "valor_bonif", "obs"]


# ── Inicialização do arquivo ──────────────────────────────────────────────────
def _init_arquivo():
    """Cria o arquivo Excel com as abas necessárias se não existir."""
    os.makedirs(DADOS_DIR, exist_ok=True)
    if os.path.exists(ARQUIVO):
        return
    wb = Workbook()
    _criar_aba(wb, "empresas",     H_EMPRESAS)
    _criar_aba(wb, "debitos",      H_DEBITOS)
    _criar_aba(wb, "bonificacoes", H_BONIFICACOES)
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.save(ARQUIVO)


def _criar_aba(wb, nome, headers):
    ws = wb.create_sheet(nome)
    ws.append(headers)
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", start_color="1F4E79")
        cell.alignment = Alignment(horizontal="center")


def _wb():
    _init_arquivo()
    return load_workbook(ARQUIVO)


def _salvar(wb):
    wb.save(ARQUIVO)


def _sheet_to_list(ws):
    """Converte uma aba em lista de dicts usando a primeira linha como chave."""
    rows = list(ws.values)
    if not rows:
        return []
    headers = [str(h) for h in rows[0]]
    return [dict(zip(headers, row)) for row in rows[1:] if any(v is not None for v in row)]


# ── Empresas ──────────────────────────────────────────────────────────────────
def listar_empresas():
    wb = _wb()
    return _sheet_to_list(wb["empresas"])


def buscar_empresa(cnpj):
    for e in listar_empresas():
        if e["cnpj"] == cnpj:
            return e
    return None


def adicionar_empresa(cnpj, razao_social):
    cnpj = cnpj.strip()
    if buscar_empresa(cnpj):
        return False, "CNPJ já cadastrado."
    wb = _wb()
    wb["empresas"].append([cnpj, razao_social.strip()])
    _salvar(wb)
    return True, "Empresa cadastrada com sucesso."


def excluir_empresa(cnpj):
    wb = _wb()
    ws = wb["empresas"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == cnpj:
            ws.delete_rows(row[0].row)
            _salvar(wb)
            return True, "Empresa removida."
    return False, "Empresa não encontrada."


# ── Débitos ───────────────────────────────────────────────────────────────────
def listar_debitos(cnpj=None):
    wb = _wb()
    dados = _sheet_to_list(wb["debitos"])
    if cnpj:
        dados = [d for d in dados if d["cnpj"] == cnpj]
    return dados


def adicionar_debito_vencimento(cnpj, nf_numero, valor_total, obs=""):
    """
    Vencimento: lançado por NF, com valor total direto (já vem do ERP).
    """
    empresa = buscar_empresa(cnpj)
    if not empresa:
        return False, "Empresa não encontrada."
    if not nf_numero.strip():
        return False, "Número da NF é obrigatório."
    try:
        v_total = round(float(str(valor_total).replace(",", ".")), 2)
    except ValueError:
        return False, "Valor total inválido."

    wb = _wb()
    wb["debitos"].append([
        str(uuid.uuid4())[:8],
        datetime.now().strftime("%d/%m/%Y %H:%M"),
        cnpj,
        empresa["razao_social"],
        "vencimento",
        nf_numero.strip(),
        None,   # produto — não se aplica
        None,   # quantidade — não se aplica
        None,   # valor_unit — não se aplica
        v_total,
        obs.strip(),
    ])
    _salvar(wb)
    return True, f"Vencimento NF {nf_numero} de R$ {v_total:.2f} registrado."


def adicionar_debito_rebaxa(cnpj, produto, quantidade, valor_unit, obs=""):
    """
    Rebaxa: registrada por produto com quantidade e diferença unitária.
    valor_unit = diferença entre preço original e preço rebaixado (ex: 0,50)
    """
    empresa = buscar_empresa(cnpj)
    if not empresa:
        return False, "Empresa não encontrada."
    try:
        qtd      = float(str(quantidade).replace(",", "."))
        v_unit   = float(str(valor_unit).replace(",", "."))
        v_debito = round(qtd * v_unit, 2)
    except ValueError:
        return False, "Quantidade ou valor inválido."

    wb = _wb()
    wb["debitos"].append([
        str(uuid.uuid4())[:8],
        datetime.now().strftime("%d/%m/%Y %H:%M"),
        cnpj,
        empresa["razao_social"],
        "rebaxa",
        None,          # nf_numero — não se aplica
        produto.strip(),
        qtd,
        v_unit,
        v_debito,
        obs.strip(),
    ])
    _salvar(wb)
    return True, f"Rebaxa de R$ {v_debito:.2f} registrada."


def excluir_debito(id_debito):
    wb = _wb()
    ws = wb["debitos"]
    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) == id_debito:
            ws.delete_rows(row[0].row)
            _salvar(wb)
            return True, "Débito removido."
    return False, "Débito não encontrado."


# ── Bonificações ──────────────────────────────────────────────────────────────
def listar_bonificacoes(cnpj=None):
    wb = _wb()
    dados = _sheet_to_list(wb["bonificacoes"])
    if cnpj:
        dados = [b for b in dados if b["cnpj"] == cnpj]
    return dados


def adicionar_bonificacao(cnpj, nf_numero, valor_total, obs=""):
    """
    Bonificação: registrada por NF com valor total.
    O valor abate diretamente no saldo devedor da empresa.
    """
    empresa = buscar_empresa(cnpj)
    if not empresa:
        return False, "Empresa não encontrada."
    if not nf_numero.strip():
        return False, "Número da NF é obrigatório."
    try:
        v_total = round(float(str(valor_total).replace(",", ".")), 2)
    except ValueError:
        return False, "Valor total inválido."

    wb = _wb()
    wb["bonificacoes"].append([
        str(uuid.uuid4())[:8],
        datetime.now().strftime("%d/%m/%Y %H:%M"),
        cnpj,
        empresa["razao_social"],
        nf_numero.strip(),
        None,    # produto — não se aplica
        None,    # quantidade — não se aplica
        None,    # valor_unit — não se aplica
        v_total,
        obs.strip(),
    ])
    _salvar(wb)
    return True, f"Bonificação NF {nf_numero} de R$ {v_total:.2f} registrada."


def excluir_bonificacao(id_bonif):
    wb = _wb()
    ws = wb["bonificacoes"]
    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) == id_bonif:
            ws.delete_rows(row[0].row)
            _salvar(wb)
            return True, "Bonificação removida."
    return False, "Bonificação não encontrada."


# ── Saldo por empresa ─────────────────────────────────────────────────────────
def calcular_saldo(cnpj):
    debitos      = listar_debitos(cnpj)
    bonificacoes = listar_bonificacoes(cnpj)
    total_deb  = sum(float(d["valor_debito"] or 0) for d in debitos)
    total_bon  = sum(float(b["valor_bonif"]  or 0) for b in bonificacoes)
    return {
        "total_debito":      round(total_deb, 2),
        "total_bonificacao": round(total_bon, 2),
        "saldo_devedor":     round(total_deb - total_bon, 2),
    }


def resumo_empresas():
    """Retorna todas as empresas com seus respectivos saldos."""
    resultado = []
    for emp in listar_empresas():
        saldo = calcular_saldo(emp["cnpj"])
        resultado.append({**emp, **saldo})
    return resultado
